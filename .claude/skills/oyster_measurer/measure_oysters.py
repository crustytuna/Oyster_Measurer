"""
measure_oysters.py — Oyster Dimension Measurement Tool
Pacific Oyster Lab — Goose Point

Pipeline:
  1. Load image
  2. Calibrate: detect ruler tick marks → compute px/mm ratio
  3. Detect oysters:
       - If a blue-mask image is provided: extract blue-painted regions (HSV)
       - Otherwise: adaptive thresholding + watershed on the raw image
  4. Measure each oyster: PCA axes → length (major) & width (minor)
  5. Export xlsx matching the reference data format
  6. Save annotated diagnostic images showing every step

Usage:
  python3 measure_oysters.py <image_path> [output_dir] [site] [initials] [mask_path]

  mask_path: optional path to a blue-painted mask image where each oyster
             is coloured solid blue. If omitted, falls back to adaptive threshold.
"""

import cv2
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import sys
import os
import re

# ── Configuration ─────────────────────────────────────────────────────────────
RULER_KNOWN_MM   = 10          # mm between major (cm) graduation marks on the ruler
MIN_OYSTER_PX    = 2_000       # minimum contour area (pixels²) to count as oyster
MAX_OYSTER_PX    = 500_000     # maximum (avoids picking up the whole table)
RULER_ROI_FRAC   = (0.818, 0.52, 0.836, 0.64)  # fallback ROI when no mask red-box is available

# ── Colours (BGR for OpenCV, then converted) ──────────────────────────────────
COL_LENGTH = (0,   220,  50)   # green  (BGR → RGB: 50,220,0)
COL_WIDTH  = (255, 120,  50)   # blue   (BGR → RGB: 50,120,255)
COL_CENTER = (255,  50,  50)   # red    (BGR → RGB: 50,50,255)
COL_RULER  = (0,  200, 255)    # yellow-ish

# ─────────────────────────────────────────────────────────────────────────────
def load_image(path):
    img = cv2.imread(str(path))
    if img is None:
        raise FileNotFoundError(f"Cannot open image: {path}")
    return img

# ── STEP 1: Ruler calibration ─────────────────────────────────────────────────
def calibrate_ruler(img, known_mm=RULER_KNOWN_MM, roi_frac=RULER_ROI_FRAC):
    """
    Detect the ruler's major graduation marks (cm lines) in the ROI and compute
    px/mm = median_spacing_between_major_marks / known_mm.

    Strategy:
    1. Try to detect long dark vertical lines (tick marks) via Hough lines
       within the ruler ROI.
    2. Fall back to column-projection peak detection if Hough yields < 3 lines.
    3. The detected peak spacing represents 'known_mm' millimetres, so
       px_per_mm = spacing / known_mm.
    """
    from scipy.signal import find_peaks, savgol_filter

    h, w = img.shape[:2]
    x0 = int(roi_frac[0] * w);  y0 = int(roi_frac[1] * h)
    x1 = int(roi_frac[2] * w);  y1 = int(roi_frac[3] * h)

    roi  = img[y0:y1, x0:x1]
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    roi_h, roi_w = gray.shape

    # ── Try Hough line detection for major tick marks ─────────────────────────
    edges = cv2.Canny(gray, 30, 100)
    # Detect lines that span at least 30% of the ROI height (major ticks only)
    min_line_len = int(roi_h * 0.30)
    lines = cv2.HoughLinesP(edges, rho=1, theta=np.pi/2,
                            threshold=15, minLineLength=min_line_len, maxLineGap=4)

    hough_x = []
    if lines is not None:
        for ln in lines:
            x1l, y1l, x2l, y2l = ln[0]
            # Only nearly-vertical lines
            if abs(x2l - x1l) < 5:
                hough_x.append((x1l + x2l) // 2)

    hough_x = sorted(set(hough_x))
    # Cluster nearby detections (within 8 px) into single marks
    clustered = []
    for xv in hough_x:
        if not clustered or xv - clustered[-1] > 8:
            clustered.append(xv)
    hough_x = clustered

    # ── Column projection fallback ────────────────────────────────────────────
    # Invert: on a white/light ruler, tick marks are dark → after inversion, bright
    inv = cv2.bitwise_not(gray)
    # Only keep pixels darker than the 40th percentile (tick marks)
    _, thresh = cv2.threshold(gray, int(np.percentile(gray, 40)), 255,
                               cv2.THRESH_BINARY_INV)
    col_proj = thresh.sum(axis=0).astype(float)
    wl = max(5, (roi_w // 30) | 1)  # window must be odd
    smooth = savgol_filter(col_proj, window_length=wl, polyorder=2)

    # Look for peaks with minimum distance matching expected cm spacing
    # Expected: ruler is maybe 150mm long in ~roi_w pixels → 1mm ≈ roi_w/150
    # cm marks are 10mm apart → expected distance ≈ roi_w/15
    min_dist = max(8, roi_w // 20)
    peaks, _ = find_peaks(smooth,
                          distance=min_dist,
                          prominence=smooth.max() * 0.08)

    # Choose best source: Hough (if ≥3 marks found) else projection peaks
    if len(hough_x) >= 3:
        spacings   = np.diff(hough_x)
        use_peaks  = np.array(hough_x)
        source     = "Hough lines"
    elif len(peaks) >= 2:
        spacings   = np.diff(peaks)
        use_peaks  = peaks
        source     = "column projection"
    else:
        raise RuntimeError(
            "Ruler calibration failed: could not detect graduation marks.\n"
            "Adjust RULER_ROI_FRAC to better frame the ruler, or set "
            "MANUAL_PX_PER_MM in the script.")

    # Filter out outlier spacings (keep within 50% of median)
    med = np.median(spacings)
    good = spacings[np.abs(spacings - med) < 0.5 * med]
    if len(good) == 0:
        good = spacings
    median_spacing = float(np.median(good))

    # Each detected mark is 'known_mm' apart → divide to get px per mm
    px_per_mm = median_spacing / known_mm

    print(f"   Calibration source: {source}")
    print(f"   Marks detected: {len(use_peaks)}  |  "
          f"median mark spacing: {median_spacing:.1f} px  |  "
          f"assumed interval: {known_mm} mm")

    tick_x_full = use_peaks + x0
    tick_y_full = (y0 + y1) // 2

    return px_per_mm, (x0, y0, x1, y1), tick_x_full, tick_y_full, col_proj, smooth, peaks

# ── Red-box calibration (preferred when a mask image is available) ────────────
# The user draws a red rectangle around the caliper/ruler in the masked PNG.
# This function finds that box, crops the raw image to it, and detects the
# tick or checker spacing — the same approach used for batch images 21-50.

RED_CAL_LOWER1 = np.array([0,   120, 100]); RED_CAL_UPPER1 = np.array([8,   255, 255])
RED_CAL_LOWER2 = np.array([168, 120, 100]); RED_CAL_UPPER2 = np.array([180, 255, 255])

def calibrate_from_red_box(raw_img, mask_img):
    """
    Detect the red box drawn in mask_img, crop raw_img to that region,
    then find the tick-mark (or checker-square) spacing to compute px/mm.

    Returns (px_per_mm, method_str, ruler_roi) or (None, reason, None).

    Three detection strategies (tried in order):
      1. Tick-mark projection — find the most regular peak spacing across
         multiple minimum-distance thresholds; divide by 10 (10 mm per major tick).
      2. Checker pattern — same projection but the dominant spacing is one
         square width (also 10 mm).
      3. Silver-body extent fallback — span of the metallic caliper body / 150 mm.
    """
    from scipy.signal import find_peaks, savgol_filter

    hsv = cv2.cvtColor(mask_img, cv2.COLOR_BGR2HSV)
    red = cv2.bitwise_or(
        cv2.inRange(hsv, RED_CAL_LOWER1, RED_CAL_UPPER1),
        cv2.inRange(hsv, RED_CAL_LOWER2, RED_CAL_UPPER2),
    )
    k = cv2.getStructuringElement(cv2.MORPH_RECT, (20, 20))
    red = cv2.morphologyEx(red, cv2.MORPH_CLOSE, k, iterations=2)
    cnts, _ = cv2.findContours(red, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = [c for c in cnts if cv2.contourArea(c) > 1000]
    if not cnts:
        return None, "no red box found in mask", None

    x0, y0, bw, bh = cv2.boundingRect(max(cnts, key=cv2.contourArea))
    ruler_roi = (x0, y0, x0 + bw, y0 + bh)
    crop = raw_img[y0:y0 + bh, x0:x0 + bw]
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)

    best = None
    for axis, proj in [("h", gray.mean(axis=0)), ("v", gray.mean(axis=1))]:
        length = len(proj)
        if length < 20:
            continue
        wl = min(max(length // 20 | 1, 5), 51)
        if wl % 2 == 0:
            wl += 1
        smooth = savgol_filter(proj, wl, 3)
        for inv in [True, False]:
            sig = -smooth if inv else smooth
            for min_d in [6, 12, 20, 35]:
                peaks, _ = find_peaks(sig, distance=min_d, prominence=2)
                if len(peaks) < 3:
                    continue
                spacings = np.diff(peaks)
                med = float(np.median(spacings))
                cv_val = np.std(spacings) / med if med > 0 else 99
                score = len(peaks) * 2 - cv_val * 15
                if best is None or score > best[0]:
                    best = (score, med, len(peaks), axis)

    if best is not None:
        px_per_mm = round(best[1] / 10.0, 2)
        return px_per_mm, f"red-box tick detection ({best[3]}-axis, {best[2]} peaks)", ruler_roi

    # Fallback: span of the silver caliper body ≈ 150 mm
    hsv2 = cv2.cvtColor(crop, cv2.COLOR_BGR2HSV)
    silver = cv2.inRange(hsv2, np.array([0, 0, 100]), np.array([180, 60, 255]))
    ch, cw = crop.shape[:2]
    proj = silver.mean(axis=0) if cw >= ch else silver.mean(axis=1)
    occupied = np.where(proj > 30)[0]
    if len(occupied) < 10:
        return None, "calibration failed (no ticks and no silver body detected)", None
    span = float(occupied[-1] - occupied[0])
    return round(span / 150.0, 2), "silver body / 150 mm", ruler_roi

# ── STEP 1b: YOLO-based caliper detection (mask-free calibration) ─────────────
_CALIPER_MODEL_PATH = Path(__file__).parent / "models" / "caliper_model.pt"
_caliper_model = None

def _get_caliper_model():
    global _caliper_model
    if _caliper_model is not None:
        return _caliper_model
    if not _CALIPER_MODEL_PATH.exists():
        return None
    try:
        from ultralytics import YOLO as _YOLO
        _caliper_model = _YOLO(str(_CALIPER_MODEL_PATH))
        return _caliper_model
    except Exception:
        return None

def calibrate_from_yolo(raw_img):
    """
    Detect the caliper in raw_img using the trained YOLO detection model,
    crop that region, then run tick-mark detection to compute px/mm.
    Falls back to None if the model is unavailable or detection fails.
    Returns (px_per_mm, method_str, ruler_roi) or (None, reason, None).
    """
    model = _get_caliper_model()
    if model is None:
        return None, "caliper_model.pt not available", None

    results = model.predict(raw_img, conf=0.3, verbose=False)
    boxes = results[0].boxes
    if not len(boxes):
        return None, "caliper YOLO found no caliper in image", None

    # Use the highest-confidence detection
    best_idx = int(boxes.conf.argmax())
    x1, y1, x2, y2 = map(int, boxes.xyxy[best_idx].cpu().numpy())
    # Add a small margin so ticks at the edges aren't clipped
    h, w = raw_img.shape[:2]
    pad = 20
    x1, y1 = max(0, x1-pad), max(0, y1-pad)
    x2, y2 = min(w, x2+pad), min(h, y2+pad)
    ruler_roi = (x1, y1, x2, y2)

    crop = raw_img[y1:y2, x1:x2]
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)

    from scipy.signal import find_peaks, savgol_filter
    best = None
    for axis, proj in [("h", gray.mean(axis=0)), ("v", gray.mean(axis=1))]:
        length = len(proj)
        if length < 20:
            continue
        wl = min(max(length // 20 | 1, 5), 51)
        if wl % 2 == 0:
            wl += 1
        smooth = savgol_filter(proj, wl, 3)
        for inv in [True, False]:
            sig = -smooth if inv else smooth
            for min_d in [6, 12, 20, 35]:
                peaks, _ = find_peaks(sig, distance=min_d, prominence=2)
                if len(peaks) < 3:
                    continue
                spacings = np.diff(peaks)
                med = float(np.median(spacings))
                cv_val = np.std(spacings) / med if med > 0 else 99
                score = len(peaks) * 2 - cv_val * 15
                if best is None or score > best[0]:
                    best = (score, med, len(peaks), axis)

    if best is not None:
        conf = float(boxes.conf[best_idx])
        px_per_mm = round(best[1] / 10.0, 2)
        return px_per_mm, f"YOLO caliper + tick detection ({best[3]}-axis, {best[2]} peaks, det_conf={conf:.2f})", ruler_roi

    # Fallback: silver body extent
    hsv2 = cv2.cvtColor(crop, cv2.COLOR_BGR2HSV)
    silver = cv2.inRange(hsv2, np.array([0, 0, 100]), np.array([180, 60, 255]))
    ch, cw = crop.shape[:2]
    proj = silver.mean(axis=0) if cw >= ch else silver.mean(axis=1)
    occupied = np.where(proj > 30)[0]
    if len(occupied) < 10:
        return None, "caliper detected but tick/body detection failed", None
    span = float(occupied[-1] - occupied[0])
    return round(span / 150.0, 2), "YOLO caliper + silver body / 150 mm", ruler_roi


# ── STEP 2a: YOLO model segmentation (primary, mask-free) ────────────────────
# Model lives next to this script; absent = fall back to blue mask or threshold.
_YOLO_MODEL_PATH = Path(__file__).parent / "models" / "oyster_model.pt"
_yolo_model = None  # loaded lazily on first call

def _get_yolo_model():
    global _yolo_model
    if _yolo_model is None:
        try:
            from ultralytics import YOLO as _YOLO
        except ImportError:
            print(
                "\n  [YOLO] ultralytics is not installed — cannot use trained model.\n"
                "  Install it with:  pip install ultralytics\n"
                "  Falling back to adaptive-threshold detection (less accurate).\n"
            )
            _yolo_model = False
            return None

        if not _YOLO_MODEL_PATH.exists():
            print(
                f"\n  [YOLO] Model file not found: {_YOLO_MODEL_PATH}\n"
                f"  Download oyster_model.pt from the repo and place it next to this script.\n"
                f"  Falling back to adaptive-threshold detection (less accurate).\n"
            )
            _yolo_model = False
            return None

        try:
            _yolo_model = _YOLO(str(_YOLO_MODEL_PATH))
        except Exception as exc:
            print(
                f"\n  [YOLO] Failed to load model ({_YOLO_MODEL_PATH.name}): {exc}\n"
                f"  The file may be corrupted or from an incompatible ultralytics version.\n"
                f"  Falling back to adaptive-threshold detection (less accurate).\n"
            )
            _yolo_model = False
            return None

    return _yolo_model if _yolo_model else None

def segment_from_yolo(img):
    """
    Run the trained YOLOv8-seg model on a raw image. Returns a list of contours
    (one per detected oyster), or None if the model is unavailable.
    """
    model = _get_yolo_model()
    if model is None:
        return None

    h, w = img.shape[:2]
    scale = 1024 / max(h, w)
    img_small = cv2.resize(img, (int(w * scale), int(h * scale)))

    results = model.predict(
        img_small,
        conf=0.25,
        iou=0.4,
        imgsz=1024,
        device="mps" if _mps_available() else "cpu",
        verbose=False,
        max_det=500,
    )

    r = results[0]
    if r.masks is None or len(r.masks) == 0:
        return []

    contours = []
    for mask_data in r.masks.xy:
        pts = (mask_data / scale).astype(np.int32)  # scale back to original size
        if len(pts) < 3:
            continue
        c = pts.reshape(-1, 1, 2)
        if MIN_OYSTER_PX < cv2.contourArea(c) < MAX_OYSTER_PX:
            contours.append(c)

    def sort_key(c):
        M = cv2.moments(c)
        if M["m00"] == 0:
            return (0, 0)
        return (int(M["m01"] / M["m00"]) // 80, int(M["m10"] / M["m00"]))

    contours.sort(key=sort_key)
    return contours

def _mps_available():
    try:
        import torch
        return torch.backends.mps.is_available()
    except Exception:
        return False

def yolo_model_available():
    return _YOLO_MODEL_PATH.exists()

# ── STEP 2b: Segmentation from blue-painted mask ──────────────────────────────
# Blue HSV range (OpenCV scale H: 0–180)
BLUE_HSV_LOWER = np.array([100,  70,  50])
BLUE_HSV_UPPER = np.array([130, 255, 255])

def segment_from_blue_mask(img, mask_img):
    """
    Detect blue-painted oyster regions in mask_img, then apply watershed to
    separate touching individuals. Returns (contours, raw_blue_binary_mask).

    Each connected blob is processed independently so the distance-transform
    threshold (40% of local max) is not dominated by a single large oyster,
    which previously caused small neighbours to lose their watershed seed and
    merge into the large one.
    """
    hsv = cv2.cvtColor(mask_img, cv2.COLOR_BGR2HSV)
    blue_binary = cv2.inRange(hsv, BLUE_HSV_LOWER, BLUE_HSV_UPPER)

    # Only remove tiny speckles — NO closing, which would bridge intentional gaps
    # between oysters that the user painted with deliberate space between them.
    open_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    cleaned = cv2.morphologyEx(blue_binary, cv2.MORPH_OPEN, open_kernel, iterations=1)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    contours = []

    # Split into individual blobs so each gets its own local distance threshold
    num_labels, labels_map = cv2.connectedComponents(cleaned, connectivity=8)

    ws_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    for blob_id in range(1, num_labels):
        blob_mask = np.where(labels_map == blob_id, np.uint8(255), np.uint8(0))

        dist = cv2.distanceTransform(blob_mask, cv2.DIST_L2, 5)
        if dist.max() == 0:
            continue

        # 40% of THIS blob's peak — not the global image max
        _, sure_fg = cv2.threshold(dist, 0.40 * dist.max(), 255, 0)
        sure_fg = sure_fg.astype(np.uint8)

        sure_bg = cv2.dilate(blob_mask, ws_kernel, iterations=3)
        unknown = cv2.subtract(sure_bg, sure_fg)

        _, markers = cv2.connectedComponents(sure_fg)
        markers += 1
        markers[unknown == 255] = 0

        img_ws = img.copy()
        markers = cv2.watershed(img_ws, markers)

        for label in np.unique(markers):
            if label <= 1:
                continue
            m = np.zeros(gray.shape, np.uint8)
            m[markers == label] = 255
            cnts, _ = cv2.findContours(m, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for c in cnts:
                if MIN_OYSTER_PX < cv2.contourArea(c) < MAX_OYSTER_PX:
                    contours.append(c)

    def sort_key(c):
        M = cv2.moments(c)
        if M["m00"] == 0:
            return (0, 0)
        return (int(M["m01"] / M["m00"]) // 80, int(M["m10"] / M["m00"]))

    contours.sort(key=sort_key)
    return contours, blue_binary

# ── STEP 2b: Fallback segmentation via adaptive threshold ─────────────────────
def segment_oysters(img):
    """
    Convert to LAB colour space, use the A channel (red-green) to
    separate brownish oysters from the white/grey table surface.
    Watershed is applied to split touching individuals.
    Returns list of contours.
    """
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    L, A, B = cv2.split(lab)

    # Oysters tend to be darker (low L) and more saturated (higher B)
    # Build a combined darkness + saturation mask
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (9, 9), 0)

    # Adaptive threshold on the L channel
    thresh_L = cv2.adaptiveThreshold(
        L, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        blockSize=51, C=8)

    # Morphological clean-up
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    cleaned = cv2.morphologyEx(thresh_L, cv2.MORPH_CLOSE, kernel, iterations=2)
    cleaned = cv2.morphologyEx(cleaned, cv2.MORPH_OPEN,  kernel, iterations=1)

    # --- Watershed to split touching oysters ---
    dist = cv2.distanceTransform(cleaned, cv2.DIST_L2, 5)
    # Higher threshold (0.55) = only accept well-separated cores as seeds
    # → fewer, larger watershed regions → less over-segmentation
    _, sure_fg = cv2.threshold(dist, 0.55 * dist.max(), 255, 0)
    sure_fg = sure_fg.astype(np.uint8)

    sure_bg = cv2.dilate(cleaned, kernel, iterations=3)
    unknown = cv2.subtract(sure_bg, sure_fg)

    _, markers = cv2.connectedComponents(sure_fg)
    markers += 1
    markers[unknown == 255] = 0

    img_ws = img.copy()
    markers = cv2.watershed(img_ws, markers)

    # Extract contours from each watershed region
    contours = []
    for label in np.unique(markers):
        if label <= 1:   # background or border
            continue
        mask = np.zeros(gray.shape, np.uint8)
        mask[markers == label] = 255
        cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for c in cnts:
            area = cv2.contourArea(c)
            if MIN_OYSTER_PX < area < MAX_OYSTER_PX:
                contours.append(c)

    # Sort left-to-right, top-to-bottom (reading order)
    def sort_key(c):
        M = cv2.moments(c)
        if M["m00"] == 0:
            return (0, 0)
        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])
        return (cy // 80, cx)   # row-major buckets of 80 px

    contours.sort(key=sort_key)
    return contours

# ── STEP 3: Measure one oyster via contour projection ────────────────────────
def measure_oyster(contour):
    """
    Fit an ellipse to the contour (cv2.fitEllipse).
    Length = major axis diameter, width = minor axis diameter.
    Ellipse fitting averages over all contour points so bumps are smoothed out.
    Falls back to minAreaRect for contours with fewer than 5 points.
    Returns (cx, cy, length_px, width_px, angle_deg, eigvec)
    """
    M = cv2.moments(contour)
    cx = float(M["m10"] / M["m00"]) if M["m00"] else 0.0
    cy = float(M["m01"] / M["m00"]) if M["m00"] else 0.0

    pts = contour.reshape(-1, 2)
    if len(pts) >= 5:
        (ecx, ecy), (minor_ax, major_ax), angle = cv2.fitEllipse(contour)
        length_px = float(max(major_ax, minor_ax))
        width_px  = float(min(major_ax, minor_ax))
        # fitEllipse angle is along axes[0] (minor_ax here); rotate if major is longer
        long_angle = angle if minor_ax >= major_ax else angle + 90.0
    else:
        (_, _), (rw, rh), rect_angle = cv2.minAreaRect(contour)
        length_px  = float(max(rw, rh))
        width_px   = float(min(rw, rh))
        long_angle = rect_angle if rw >= rh else rect_angle + 90.0

    angle_rad = np.radians(long_angle)
    ev0 = np.array([ np.cos(angle_rad),  np.sin(angle_rad)], dtype=np.float32)
    ev1 = np.array([-np.sin(angle_rad),  np.cos(angle_rad)], dtype=np.float32)
    eigvec = np.array([ev0, ev1])

    return float(cx), float(cy), length_px, width_px, float(long_angle), eigvec

# ── STEP 4: Draw measurement lines on a copy of the image ─────────────────────
def draw_measurements(img, contours, measurements, px_per_mm):
    vis = img.copy()
    for idx, (contour, (cx, cy, lpx, wpx, angle, eigvec)) in \
            enumerate(zip(contours, measurements), start=1):

        # Build a filled mask so lines are clipped inside the contour
        mask = np.zeros(vis.shape[:2], dtype=np.uint8)
        cv2.drawContours(mask, [contour], -1, 255, cv2.FILLED)

        half_l, half_w = lpx / 2, wpx / 2
        ev0, ev1 = eigvec[0], eigvec[1]
        p1 = (int(cx + half_l * ev0[0]), int(cy + half_l * ev0[1]))
        p2 = (int(cx - half_l * ev0[0]), int(cy - half_l * ev0[1]))
        p3 = (int(cx + half_w * ev1[0]), int(cy + half_w * ev1[1]))
        p4 = (int(cx - half_w * ev1[0]), int(cy - half_w * ev1[1]))

        line_layer = vis.copy()
        cv2.line(line_layer, p1, p2, COL_LENGTH, 5)
        cv2.line(line_layer, p3, p4, COL_WIDTH, 5)
        vis[mask > 0] = line_layer[mask > 0]

        # Red outline on top
        cv2.drawContours(vis, [contour], -1, (0, 0, 255), 1)

        # Centre dot
        cv2.circle(vis, (int(cx), int(cy)), 8, COL_CENTER, -1)

        # White number
        label = str(idx)
        font = cv2.FONT_HERSHEY_SIMPLEX
        fs = max(0.9, min(1.3, lpx / 170))
        (tw, th), _ = cv2.getTextSize(label, font, fs, 2)
        tx, ty = int(cx - tw / 2), int(cy + th / 2)
        cv2.putText(vis, label, (tx, ty), font, fs, (0, 0, 0),      5, cv2.LINE_AA)
        cv2.putText(vis, label, (tx, ty), font, fs, (255, 255, 255), 2, cv2.LINE_AA)
    return vis

# ── STEP 5: Ruler diagnostic ───────────────────────────────────────────────────
def draw_ruler_calibration(img, ruler_roi, tick_x_full, tick_y, px_per_mm):
    vis = img.copy()
    x0, y0, x1, y1 = ruler_roi
    cv2.rectangle(vis, (x0, y0), (x1, y1), COL_RULER, 3)
    # Mark detected tick positions if any were found
    for tx in tick_x_full[:20]:
        cv2.line(vis, (tx, tick_y - 18), (tx, tick_y + 18), COL_RULER, 2)
    # Draw a 10 mm scale bar anchored to the ruler ROI
    bar_x = x0 + 10
    bar_y = y1 + 30
    bar_end = bar_x + int(10 * px_per_mm)
    cv2.line(vis, (bar_x, bar_y), (bar_end, bar_y), COL_RULER, 3)
    cv2.line(vis, (bar_x, bar_y - 8), (bar_x, bar_y + 8), COL_RULER, 2)
    cv2.line(vis, (bar_end, bar_y - 8), (bar_end, bar_y + 8), COL_RULER, 2)
    cv2.putText(vis, f"10 mm = {int(10*px_per_mm)} px",
                (bar_x, bar_y + 28),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, COL_RULER, 2)
    cv2.putText(vis, f"{px_per_mm:.2f} px/mm",
                (x0, y0 - 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, COL_RULER, 2)
    return vis

# ── STEP 6: Export xlsx ────────────────────────────────────────────────────────
def export_xlsx(measurements, px_per_mm, image_path, out_path,
                site="Goose Point", initials="CC"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Parse date + tag from filename  e.g. 20260522_bag380_raw.jpeg
    stem = Path(image_path).stem
    date_match = re.search(r'(\d{8})', stem)
    tag_match  = re.search(r'bag(\d+)', stem)
    image_date = int(date_match.group(1)) if date_match else 0
    tag_id     = int(tag_match.group(1))  if tag_match  else 0
    image_name = stem.replace("_raw", "_annotated")

    # Header row
    headers = ["Site","Image Date","Initials","Image Name",
               "Tag ID","Oyster","Measurement","Value mm ","Notes "]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows — 2 rows per oyster (length then width)
    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    row_num = 2
    for idx, (_, (_, _, lpx, wpx, *_)) in enumerate(measurements, start=1):
        length_mm = round(lpx / px_per_mm, 2)
        width_mm  = round(wpx / px_per_mm, 2)
        fill = alt_fill if idx % 2 == 0 else PatternFill()

        for meas, val in [("length", length_mm), ("width ", width_mm)]:
            row_data = [site, image_date, initials, image_name,
                        tag_id, idx, meas, val, None]
            for col, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=v)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            row_num += 1

    # Column widths
    for col, width in enumerate([14,14,10,34,8,8,12,12,10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(str(out_path))
    print(f"  Saved xlsx → {out_path}")

# ── STEP 7: Multi-panel diagnostic figure ─────────────────────────────────────
def save_diagnostic(img_raw, img_ruler, img_segmented, img_measured,
                    col_proj, smooth, peaks, ruler_roi, px_per_mm,
                    measurements, out_path, blue_mask=None):
    fig = plt.figure(figsize=(22, 20))
    fig.patch.set_facecolor("#0D1B2A")

    title_kw = dict(color="white", fontsize=12, fontweight="bold", pad=8)

    def show(ax, img_bgr, title, cmap=None):
        ax.set_facecolor("#0D1B2A")
        if img_bgr.ndim == 3:
            ax.imshow(cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB))
        else:
            ax.imshow(img_bgr, cmap=cmap or "gray")
        ax.set_title(title, **title_kw)
        ax.axis("off")

    # Panel 1: raw image
    ax1 = fig.add_subplot(3, 2, 1)
    show(ax1, img_raw, "① Raw Image")

    # Panel 2: ruler calibration overlay
    ax2 = fig.add_subplot(3, 2, 2)
    show(ax2, img_ruler, f"② Ruler Calibration\n{px_per_mm:.2f} px / mm")
    # inset ruler projection
    ax_ins = ax2.inset_axes([0.0, 0.0, 1.0, 0.22])
    ax_ins.set_facecolor("#0D1B2A")
    x0, y0, x1, y1 = ruler_roi
    roi_w = x1 - x0
    ax_ins.plot(col_proj[:roi_w], color="#8FA3B1", linewidth=0.8, alpha=0.6, label="raw proj")
    ax_ins.plot(smooth[:roi_w],   color="#F1C40F", linewidth=1.2, label="smoothed")
    pk_in = peaks[peaks < roi_w].astype(int)
    ax_ins.scatter(pk_in, smooth[pk_in],
                   color="#E74C3C", s=18, zorder=5, label="ticks")
    ax_ins.set_xticks([]); ax_ins.set_yticks([])
    ax_ins.spines[:].set_visible(False)

    # Panel 3: blue mask (if provided) or adaptive threshold fallback
    ax3 = fig.add_subplot(3, 2, 3)
    if blue_mask is not None:
        show(ax3, blue_mask, "③ Blue Mask Detection\n(blue-painted oysters = white)", cmap="gray")
    else:
        gray = cv2.cvtColor(img_raw, cv2.COLOR_BGR2GRAY)
        lab  = cv2.cvtColor(img_raw, cv2.COLOR_BGR2LAB)
        L    = cv2.split(lab)[0]
        thresh = cv2.adaptiveThreshold(L, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY_INV, 51, 8)
        show(ax3, thresh, "③ Adaptive Threshold Mask\n(oysters = white)", cmap="gray")

    # Panel 4: segmented / watershed
    ax4 = fig.add_subplot(3, 2, 4)
    show(ax4, img_segmented, "④ Watershed Segmentation\n(each colour = 1 oyster)")

    # Panel 5: measurements overlay
    ax5 = fig.add_subplot(3, 2, 5)
    show(ax5, img_measured,
         "⑤ Dimension Measurement Lines\nGreen = Length  |  Blue = Width")

    # Panel 6: scatter of length vs width
    ax6 = fig.add_subplot(3, 2, 6)
    ax6.set_facecolor("#132338")
    lengths = [lpx / px_per_mm for (_, (_, _, lpx, wpx, *_)) in measurements]
    widths  = [wpx / px_per_mm for (_, (_, _, lpx, wpx, *_)) in measurements]
    sc = ax6.scatter(lengths, widths, c=range(len(lengths)),
                     cmap="plasma", s=60, edgecolors="white", linewidths=0.5, alpha=0.9)
    for i, (l, w) in enumerate(zip(lengths, widths), 1):
        ax6.annotate(str(i), (l, w), fontsize=5.5, color="white",
                     ha="center", va="bottom", xytext=(0, 4),
                     textcoords="offset points")
    ax6.set_xlabel("Length (mm)", color="white", fontsize=10)
    ax6.set_ylabel("Width (mm)",  color="white", fontsize=10)
    ax6.set_title("⑥ Length vs Width per Oyster", **title_kw)
    ax6.tick_params(colors="white")
    for s in ax6.spines.values(): s.set_color("#3D5A73")
    plt.colorbar(sc, ax=ax6, label="Oyster #").ax.yaxis.label.set_color("white")

    # Legend patches
    patches = [
        mpatches.Patch(color="#00DC32", label="Length line (major PCA axis)"),
        mpatches.Patch(color="#3278FF", label="Width line (minor PCA axis)"),
        mpatches.Patch(color="#FF3232", label="Centre point"),
    ]
    fig.legend(handles=patches, loc="lower center", ncol=3,
               framealpha=0.2, labelcolor="white", fontsize=9,
               facecolor="#0D1B2A", edgecolor="#3D5A73")

    fig.suptitle(
        f"Oyster Measurement Diagnostic  ·  {Path(img_measured).name if hasattr(img_measured,'__fspath__') else 'Bag 380'}\n"
        f"Calibration: {px_per_mm:.2f} px/mm  ·  Oysters detected: {len(measurements)}",
        color="white", fontsize=14, fontweight="bold", y=0.995
    )
    plt.tight_layout(rect=[0, 0.04, 1, 0.99])
    plt.savefig(str(out_path), dpi=160, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close()
    print(f"  Saved diagnostic → {out_path}")

# ── Colour-coded segmentation visualisation ───────────────────────────────────
def coloured_segments(img, contours):
    vis = img.copy()
    np.random.seed(42)
    for c in contours:
        col = [int(x) for x in np.random.randint(80, 255, 3)]
        cv2.drawContours(vis, [c], -1, col, cv2.FILLED)
    blended = cv2.addWeighted(img, 0.45, vis, 0.55, 0)
    return blended

# ── Main ──────────────────────────────────────────────────────────────────────
def run(image_path, out_dir=None, site="Goose Point", initials="CC", mask_path=None, px_per_mm_override=None):
    image_path = Path(image_path)
    if out_dir is None:
        out_dir = Path.home() / "Desktop"
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    stem = image_path.stem.replace("_raw", "")
    print(f"\n{'='*60}")
    print(f"  Oyster Measurer — {image_path.name}")
    print(f"{'='*60}")

    # ① Load
    print("\n① Loading image...")
    img = load_image(image_path)
    h, w = img.shape[:2]
    print(f"   {w} × {h} px")

    # ② Calibrate
    print("\n② Calibrating ruler...")
    tick_x = np.array([], dtype=int)
    col_proj = np.zeros(10); smooth_proj = np.zeros(10); peaks = np.array([])

    if px_per_mm_override is not None:
        # Explicit user-supplied value — loud, not silent
        px_per_mm = float(px_per_mm_override)
        h, w = img.shape[:2]
        rf = RULER_ROI_FRAC
        ruler_roi = (int(rf[0]*w), int(rf[1]*h), int(rf[2]*w), int(rf[3]*h))
        tick_y = (ruler_roi[1] + ruler_roi[3]) // 2
        print(f"   Manual override: {px_per_mm:.3f} px/mm  ← measurements will be wrong if this is off")

    elif mask_path is not None and Path(mask_path).exists():
        # Preferred path: red box drawn in the mask locates the ruler automatically
        mask_img_cal = load_image(mask_path)
        px_per_mm, cal_method, red_roi = calibrate_from_red_box(img, mask_img_cal)
        if px_per_mm is None:
            raise RuntimeError(
                f"Calibration from mask red box failed: {cal_method}\n"
                f"  Draw a red rectangle around the ruler in the masked image,\n"
                f"  or pass --px-per-mm <value> as a manual override."
            )
        ruler_roi = red_roi
        tick_y = (ruler_roi[1] + ruler_roi[3]) // 2
        print(f"   Calibration: {px_per_mm:.3f} px/mm  ({cal_method})")

    else:
        # Fallback: use the hardcoded ROI constant (works when ruler is in the default position)
        print(f"   No mask provided — attempting ROI-based calibration (RULER_ROI_FRAC={RULER_ROI_FRAC})")
        print(f"   If this fails or gives a wrong value, draw a red box around the ruler in a mask image.")
        px_per_mm, ruler_roi, tick_x, tick_y, col_proj, smooth_proj, peaks = \
            calibrate_ruler(img, known_mm=RULER_KNOWN_MM)
        print(f"   Calibration: {px_per_mm:.3f} px/mm  (ROI tick detection)")

    # Sanity check — hard-fail for clearly impossible values
    if px_per_mm > 50 or px_per_mm < 1.0:
        raise RuntimeError(
            f"Calibration result {px_per_mm:.2f} px/mm is implausible for a field photo "
            f"(expected 3–25 px/mm).\n"
            f"  Check that the ruler is visible, adjust RULER_ROI_FRAC, or pass --px-per-mm."
        )
    img_ruler = draw_ruler_calibration(img, ruler_roi, tick_x, tick_y, px_per_mm)

    # ③ Segment
    print("\n③ Segmenting oysters...")
    blue_mask_vis = None
    if mask_path is not None:
        mask_img = load_image(mask_path)
        print(f"   Using blue-mask image: {Path(mask_path).name}")
        contours, blue_binary = segment_from_blue_mask(img, mask_img)
        blue_mask_vis = blue_binary
    elif yolo_model_available():
        print("   Using trained YOLOv8 model (oyster_model.pt)")
        contours = segment_from_yolo(img)
        if contours is None:
            print("   YOLO failed — falling back to adaptive threshold")
            contours = segment_oysters(img)
    else:
        print("   No mask, no model — falling back to adaptive threshold")
        contours = segment_oysters(img)
    print(f"   Detected {len(contours)} oyster individuals")
    img_seg = coloured_segments(img, contours)

    # ④ Measure
    print("\n④ Measuring dimensions...")
    measurements = []
    for i, c in enumerate(contours, 1):
        m = measure_oyster(c)
        cx, cy, lpx, wpx, angle, _ = m
        lmm = lpx / px_per_mm
        wmm = wpx / px_per_mm
        print(f"   Oyster {i:>3d}: length={lmm:6.2f} mm  width={wmm:6.2f} mm  "
              f"(angle={angle:.1f}°)")
        measurements.append((c, m))

    img_meas = draw_measurements(img, contours,
                                  [m for (_, m) in measurements], px_per_mm)

    # ⑤ Export xlsx
    print("\n⑤ Exporting xlsx...")
    xlsx_path = out_dir / f"{stem}_measured.xlsx"
    export_xlsx(measurements, px_per_mm, image_path, xlsx_path,
                site=site, initials=initials)

    # ⑥ Save annotated image
    ann_path = out_dir / f"{stem}_annotated_measured.png"
    cv2.imwrite(str(ann_path), img_meas)
    print(f"  Saved annotated image → {ann_path}")

    # ⑦ Diagnostic figure
    print("\n⑥ Saving diagnostic figure...")
    diag_path = out_dir / f"{stem}_diagnostic.png"
    save_diagnostic(img, img_ruler, img_seg, img_meas,
                    col_proj, smooth_proj, peaks, ruler_roi, px_per_mm,
                    measurements, diag_path, blue_mask=blue_mask_vis)

    print(f"\n{'='*60}")
    print(f"  DONE  —  {len(measurements)} oysters measured")
    print(f"  xlsx       → {xlsx_path}")
    print(f"  annotated  → {ann_path}")
    print(f"  diagnostic → {diag_path}")
    print(f"{'='*60}\n")
    return measurements, px_per_mm

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(
        description="Measure Pacific oyster dimensions from an overhead field photograph."
    )
    ap.add_argument("image_path",  help="Path to the raw oyster JPEG/PNG")
    ap.add_argument("output_dir",  nargs="?", default=None,
                    help="Where to save outputs (default: ~/Desktop)")
    ap.add_argument("site",        nargs="?", default="Unknown Site",
                    help="Site name written into the xlsx")
    ap.add_argument("initials",    nargs="?", default="--",
                    help="Measurer initials written into the xlsx")
    ap.add_argument("mask_path",   nargs="?", default=None,
                    help="Blue-painted mask PNG (also used for red-box ruler calibration)")
    ap.add_argument("--px-per-mm", type=float, default=None,
                    help="Manual px/mm override — skips auto-calibration. "
                         "Use only when auto-calibration cannot find the ruler.")
    args = ap.parse_args()
    run(args.image_path, args.output_dir, args.site, args.initials,
        args.mask_path, px_per_mm_override=args.px_per_mm)
