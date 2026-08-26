"""
Oyster Measurement Web App
==========================
Streamlit interface for the oyster measurement pipeline.

Workflow:
  1. Upload one or more field photos
  2. Enter px/mm calibration for each photo (read from the in-frame caliper)
  3. Click Run Analysis — YOLOv8 segments oysters, PCA measures each one
  4. View annotated images + per-image tables, then download combined XLSX

Run:
  streamlit run app.py
"""

import io
import re
import sys
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
import openpyxl
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Import the measurement pipeline ────────────────────────────────────────────
_SKILLS_DIR = Path(__file__).parent / ".claude" / "skills" / "oyster_measurer"
sys.path.insert(0, str(_SKILLS_DIR))
import measure_oysters as mo  # noqa: E402 (path must be set first)

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Oyster Measurer",
    page_icon="🦪",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Helpers ────────────────────────────────────────────────────────────────────

def _decode(file_bytes: bytes) -> np.ndarray:
    """Decode uploaded image bytes to a BGR numpy array."""
    arr = np.frombuffer(file_bytes, np.uint8)
    return cv2.imdecode(arr, cv2.IMREAD_COLOR)


def _thumb(img_bgr: np.ndarray, max_w: int = 420) -> np.ndarray:
    """Return a width-capped RGB thumbnail for display."""
    h, w = img_bgr.shape[:2]
    scale = min(1.0, max_w / w)
    small = cv2.resize(img_bgr, (int(w * scale), int(h * scale)))
    return cv2.cvtColor(small, cv2.COLOR_BGR2RGB)


def _write_sheet(ws, measurements, px_per_mm: float,
                 filename: str, site: str, initials: str) -> None:
    """Write one worksheet matching the lab's XLSX format."""
    stem = Path(filename).stem
    date_match = re.search(r"(\d{8})", stem)
    tag_match  = re.search(r"bag(\d+)", stem)
    image_date = int(date_match.group(1)) if date_match else 0
    tag_id     = int(tag_match.group(1))  if tag_match  else 0
    image_name = stem.replace("_raw", "_annotated")

    headers = ["Site", "Image Date", "Initials", "Image Name",
               "Tag ID", "Oyster", "Measurement", "Value mm", "Notes"]
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill   = PatternFill("solid", fgColor="D9E1F2")
    center     = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = center
        cell.border = border

    row_num = 2
    for idx, (_, _, lpx, wpx, *_) in enumerate(measurements, start=1):
        length_mm = round(lpx / px_per_mm, 2)
        width_mm  = round(wpx / px_per_mm, 2)
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        for meas, val in [("length", length_mm), ("width", width_mm)]:
            row_data = [site, image_date, initials, image_name,
                        tag_id, idx, meas, val, None]
            for col, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=v)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center
            row_num += 1

    for col, width in enumerate([14, 14, 10, 34, 8, 8, 12, 12, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width


# ── UI ─────────────────────────────────────────────────────────────────────────

st.title("🦪 Oyster Measurement Tool")
st.caption("Upload field photos · auto-calibrate px/mm from caliper · download measurements as XLSX.")

# Model status banners
if mo.yolo_model_available():
    st.success("YOLOv8 oyster segmentation model ready.", icon="✅")
else:
    st.warning(
        "YOLOv8 model not found — falling back to adaptive threshold segmentation.",
        icon="⚠️",
    )

caliper_model_available = (_SKILLS_DIR / "caliper_model.pt").exists()
if caliper_model_available:
    st.success("YOLOv8 caliper detection model ready — px/mm will be auto-detected.", icon="✅")
else:
    st.warning("Caliper model not found — manual px/mm entry required.", icon="⚠️")

st.divider()

# ── Step 1: Upload ─────────────────────────────────────────────────────────────
st.header("Step 1 — Upload photos")
files = st.file_uploader(
    "Choose field photos (JPEG or PNG, one per bag/tray)",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if not files:
    st.info("Upload at least one photo to continue.")
    st.stop()

# ── Step 2: px/mm calibration ──────────────────────────────────────────────────
st.header("Step 2 — Calibration")
n_cols = min(len(files), 3)
px_mm: dict[str, float] = {}
px_mm_method: dict[str, str] = {}
file_cache: dict[str, bytes] = {}

row_files = [files[i : i + n_cols] for i in range(0, len(files), n_cols)]
for row in row_files:
    cols = st.columns(n_cols)
    for col, f in zip(cols, row):
        raw = f.read()
        file_cache[f.name] = raw
        with col:
            img_bgr = _decode(raw)
            if img_bgr is None:
                st.error(f"Cannot decode {f.name}")
                continue
            st.image(_thumb(img_bgr), caption=f.name)

            # Try auto-calibration first
            auto_ratio, auto_method, _ = mo.calibrate_from_yolo(img_bgr)
            if auto_ratio and 1.0 <= auto_ratio <= 50.0:
                px_mm[f.name] = auto_ratio
                px_mm_method[f.name] = auto_method
                st.success(f"Auto: **{auto_ratio} px/mm** ({auto_method})")
                # Allow manual override
                override = st.number_input(
                    "Override px/mm (optional)",
                    min_value=0.5, max_value=50.0,
                    value=float(auto_ratio), step=0.01, format="%.2f",
                    key=f"px_{f.name}",
                )
                px_mm[f.name] = override
            else:
                st.warning(f"Auto-calibration failed ({auto_method}) — enter manually.")
                px_mm[f.name] = st.number_input(
                    "px/mm",
                    min_value=0.5, max_value=50.0,
                    value=5.0, step=0.01, format="%.2f",
                    key=f"px_{f.name}",
                    help="Pixels per millimetre — read from the in-frame caliper",
                )
                px_mm_method[f.name] = "manual"

# ── Step 3: Metadata ───────────────────────────────────────────────────────────
st.header("Step 3 — Metadata")
c1, c2 = st.columns(2)
site     = c1.text_input("Site name", value="Unknown Site")
initials = c2.text_input("Analyst initials", value="")

# ── Step 4: Run ────────────────────────────────────────────────────────────────
st.header("Step 4 — Run analysis")

if st.button("▶  Run Analysis", type="primary", use_container_width=True):
    wb         = openpyxl.Workbook()
    first_ws   = True
    results    = []   # (name, annotated_rgb | None, meas_list, ratio)

    bar = st.progress(0.0, text="Starting…")

    for i, f in enumerate(files):
        bar.progress(i / len(files), text=f"Processing {f.name}…")

        img_bgr = _decode(file_cache[f.name])
        if img_bgr is None:
            results.append((f.name, None, [], px_mm[f.name]))
            continue

        ratio = px_mm[f.name]

        # Detect oysters (YOLO → adaptive threshold fallback)
        contours = mo.segment_from_yolo(img_bgr)
        if contours is None:
            contours = mo.segment_oysters(img_bgr)

        if not contours:
            results.append((f.name, None, [], ratio))
            continue

        meas = [mo.measure_oyster(c) for c in contours]

        # Downsample to a display-friendly width before drawing annotations so
        # that font sizes and line widths look correct in the browser.
        ANN_W = 1400
        h_img, w_img = img_bgr.shape[:2]
        ann_scale = min(1.0, ANN_W / w_img)
        if ann_scale < 1.0:
            ann_w = int(w_img * ann_scale)
            ann_h = int(h_img * ann_scale)
            img_ann = cv2.resize(img_bgr, (ann_w, ann_h))
            contours_ann = [np.round(c.astype(float) * ann_scale).astype(np.int32)
                            for c in contours]
            meas_ann = [(cx * ann_scale, cy * ann_scale,
                         lpx * ann_scale, wpx * ann_scale, a, ev * ann_scale)
                        for cx, cy, lpx, wpx, a, ev in meas]
        else:
            img_ann, contours_ann, meas_ann = img_bgr, contours, meas

        annotated_bgr = mo.draw_measurements(img_ann, contours_ann, meas_ann, ratio)
        annotated_rgb = cv2.cvtColor(annotated_bgr, cv2.COLOR_BGR2RGB)

        results.append((f.name, annotated_rgb, meas, ratio))

        # Add to workbook (one sheet per image)
        ws = wb.active if first_ws else wb.create_sheet()
        ws.title = Path(f.name).stem[:31]
        first_ws = False
        _write_sheet(ws, meas, ratio, f.name, site, initials)

    bar.progress(1.0, text="Done!")

    # ── Results display ────────────────────────────────────────────────────────
    st.divider()
    st.header("Results")

    any_detections = False
    for fname, ann_rgb, meas, ratio in results:
        st.subheader(fname)
        if ann_rgb is None or not meas:
            st.warning(f"No oysters detected in {fname}.")
            continue
        any_detections = True

        col_img, col_tbl = st.columns([3, 2])
        with col_img:
            st.image(ann_rgb, use_container_width=True)
        with col_tbl:
            table = [
                {
                    "#": j + 1,
                    "Length (mm)": round(m[2] / ratio, 2),
                    "Width (mm)":  round(m[3] / ratio, 2),
                }
                for j, m in enumerate(meas)
            ]
            st.dataframe(table, use_container_width=True, hide_index=True)
            st.caption(f"{len(meas)} oysters detected · {ratio:.2f} px/mm")

    # ── Download button ────────────────────────────────────────────────────────
    if any_detections:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname_out = f"oyster_measurements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label="📥 Download XLSX",
            data=buf.getvalue(),
            file_name=fname_out,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    else:
        st.error("No oysters were detected in any of the uploaded images.")
